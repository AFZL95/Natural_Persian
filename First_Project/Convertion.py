import openpyxl
import codecs

text = ''
wb = openpyxl.load_workbook('mes.xlsx')
# i deleted the mes.xlsx file becuz it's around 400 MB , the "converted.txt" is available in this directory
sheet = wb.get_sheet_by_name('Query')
r = sheet.max_row
for i in range(1, r+1):
    if sheet.cell(row=i, column=1).value is not None:
        text += sheet.cell(row=i, column=1).value
        text += "\n"


myfile = codecs.open("converted.txt","w","utf-8")
myfile.write(text)
myfile.close()
